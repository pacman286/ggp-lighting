"""
generate_poles_json.py
GGP Lighting Assessment Tool — Data Generator
Run this script whenever pole data in the Excel file changes.
Usage: python generate_poles_json.py
Output: data/poles.json
"""

import json
import os
from openpyxl import load_workbook

# ── Configuration ────────────────────────────────────────────────────────────

# Map each lighting type to its source Excel file.
# Add new types here as T1, T3, T5, etc. are onboarded.
EXCEL_SOURCES = {
    "T2":   "Lighting_Type_T2.xlsx",
    "T2A":  "Lighting_Type_T2A.xlsx",
    "T3":   "Lighting_Type_T3.xlsx",
    "T3A":  "Lighting_Type_T3A.xlsx",
    "T4":   "Lighting_Type_T4.xlsx",
    "T4A":  "Lighting_Type_T4A.xlsx",
    "T5":   "Lighting_Type_T5.xlsx",
    "T9":   "Lighting_Type_T9.xlsx",
    "T10":  "Lighting_Type_T10.xlsx",
    "T14":  "Lighting_Type_T14.xlsx",
    "T14A": "Lighting_Type_T14A.xlsx",
    "T24":  "Lighting_Type_T24.xlsx",
    "T24A": "Lighting_Type_T24A.xlsx",
    "T24B": "Lighting_Type_T24B.xlsx",
    "T24C": "Lighting_Type_T24C.xlsx",
}

OUTPUT_FILE = "data/poles.json"

# Lighting item type definitions.
# Fixtures are listed top-down, left-to-right within each zone.
# Add new type blocks here when onboarding additional types.
POLE_TYPE_DEFINITIONS = {
    "T2": {
        "name": "T2",
        "description": "Single-fixture Pole",
        "subType": "Structura, 18', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Fixture",
                "zone": "Top",
                "manufacturer": "WE-EF",
                "model": "661-2421 RFL540-SE Type 3",
                "description": "Pathway Light"
            }
        ]
    },
    "T2A": {
        "name": "T2A",
        "description": "Single-fixture Pole",
        "subType": "Structura, 18', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Fixture",
                "zone": "Top",
                "manufacturer": "WE-EF",
                "model": "661-2421 RFL540-SE Type 4",
                "description": "Pathway Light"
            }
        ]
    },
    "T3": {
        "name": "T3",
        "description": "Single-fixture Pathway Pole",
        "subType": "Structura, 14', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Pathway",
                "zone": "Top",
                "manufacturer": "WE-EF",
                "model": "661-1345 RFL530-LED",
                "description": "Pathway Light"
            }
        ]
    },
    "T3A": {
        "name": "T3A",
        "description": "Multi-fixture Pathway Pole (Variant A)",
        "subType": "Structura, 18', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 2,
                "label": "Bottom Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 3,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345 RFL530-LED",
                "description": "Pathway Light"
            }
        ]
    },
    "T4": {
        "name": "T4",
        "description": "Multi-fixture Pole",
        "subType": "Structura, 45', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Mid Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Bottom Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345",
                "description": "Pathway Light"
            }
        ]
    },
    "T4A": {
        "name": "T4A",
        "description": "Multi-fixture Pole (Variant A)",
        "subType": "Structura, 45', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Mid Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Mid Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Bottom Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 6,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345",
                "description": "Pathway Light"
            }
        ]
    },
    "T5": {
        "name": "T5",
        "description": "Multi-fixture Sports Court Pole",
        "subType": "Timber, 45', Round",
        "fixtures": [
            {
                "position": 1,
                "label": "Marker",
                "zone": "Top",
                "manufacturer": "Phoenix",
                "model": "VA-W-17LED-CW-FGC-G",
                "description": "Marker Light"
            },
            {
                "position": 2,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "Beacon",
                "model": "AL-D/60NB-136/4K/3X5/UNV",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "Beacon",
                "model": "AL-D/60NB-136/4K/3X5/UNV",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Bottom Left Flood",
                "zone": "Top",
                "manufacturer": "Beacon",
                "model": "AL-D/60NB-136/4K/3X5/UNV",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Bottom Right Flood",
                "zone": "Top",
                "manufacturer": "Beacon",
                "model": "AL-D/60NB-136/4K/3X5/UNV",
                "description": "LED Floodlight"
            }
        ]
    },
    "T9": {
        "name": "T9",
        "description": "Ceiling-Mount Fixture",
        "subType": "Ceiling-Mount Fixture",
        "fixtures": [
            {
                "position": 1,
                "label": "Fixture",
                "zone": "Top",
                "manufacturer": "McGraw-Edison",
                "model": "TT-D1-740-U-WQ-WM-BZ",
                "description": "Ceiling-Mount Light"
            }
        ]
    },
    "T10": {
        "name": "T10",
        "description": "Suspended Fixture",
        "subType": "Suspended Fixture",
        "fixtures": [
            {
                "position": 1,
                "label": "Fixture",
                "zone": "Top",
                "manufacturer": "WE-EF",
                "model": "661-3326 RFS530-LED",
                "description": "Suspended Light"
            }
        ]
    },
    "T14": {
        "name": "T14",
        "description": "Multi-fixture Garden Spot Pole",
        "subType": "Structura, 14', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 2,
                "label": "Top Right Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 3,
                "label": "Bottom Left Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 4,
                "label": "Bottom Right Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            }
        ]
    },
    "T14A": {
        "name": "T14A",
        "description": "Multi-fixture Spot Pole (Variant A)",
        "subType": "Structura, 14', Tapered Square",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            },
            {
                "position": 2,
                "label": "Bottom Spot",
                "zone": "Top",
                "manufacturer": "PYR",
                "model": "LED/50W/33LED",
                "description": "LED Spotlight"
            }
        ]
    },
    "T24": {
        "name": "T24",
        "description": "Multi-fixture Bike Park Pole",
        "subType": "Timber, 45', Round",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Mid Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Mid Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Bottom Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 6,
                "label": "Bottom Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 7,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345",
                "description": "Pathway Light"
            }
        ]
    },
    "T24A": {
        "name": "T24A",
        "description": "Multi-fixture Skate Park Pole (Variant A)",
        "subType": "Timber, 45', Round",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Mid Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Mid Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Bottom Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 6,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345",
                "description": "Pathway Light"
            }
        ]
    },
    "T24B": {
        "name": "T24B",
        "description": "Multi-fixture Skate Park Pole (Variant B)",
        "subType": "Timber, 45', Round",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Mid Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Mid Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Bottom Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 6,
                "label": "Bottom Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            }
        ]
    },
    "T24C": {
        "name": "T24C",
        "description": "Multi-fixture Bike/Skate Park Pole (Variant C)",
        "subType": "Timber, 45', Round",
        "fixtures": [
            {
                "position": 1,
                "label": "Top Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 2,
                "label": "Top Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 3,
                "label": "Bottom Left Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 4,
                "label": "Bottom Right Flood",
                "zone": "Top",
                "manufacturer": "North Star Lighting",
                "model": "PRO2-151-WW-W60-U-DM-V",
                "description": "LED Floodlight"
            },
            {
                "position": 5,
                "label": "Pathway",
                "zone": "Mid",
                "manufacturer": "WE-EF",
                "model": "661-1345",
                "description": "Pathway Light"
            }
        ]
    }
}

# Condition scale — shared across all pole types
CONDITION_SCALE = [
    {"value": 0, "label": "Not Assessed",     "color": "#9e9e9e"},
    {"value": 1, "label": "100% Functional",  "color": "#2e7d32"},
    {"value": 2, "label": ">50% Functional",  "color": "#7cb342"},
    {"value": 3, "label": "<50% Functional",  "color": "#ef6c00"},
    {"value": 4, "label": "Flickering",        "color": "#d84315"},
    {"value": 5, "label": "Non-Functional",    "color": "#b71c1c"}
]

# Map display configuration
MAP_CONFIG = {
    "center": [36.1196, -95.9847],
    "defaultZoom": 16,
    "minZoom": 14,
    "maxZoom": 20,
    # Static satellite image overlay — populate after obtaining image from Google Earth Pro
    # bounds format: [[south_lat, west_lng], [north_lat, east_lng]]
    "imageOverlay": {
        "enabled": False,
        "imagePath": "images/satellite/park_satellite.jpg",
        "bounds": [[36.110, -95.992], [36.130, -95.978]],
        "opacity": 1.0
    }
}

# ── Excel Processing ─────────────────────────────────────────────────────────

def parse_sheet(wb, sheet_name):
    """Read all rows from a sheet, returning list of dicts keyed by header."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]

def str_or_null(val):
    """Return stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def int_or_null(val):
    """Return int or None."""
    try:
        return int(val)
    except (TypeError, ValueError):
        return None

def extract_coords(row):
    """
    Extract lat/lng robustly regardless of column label accuracy.
    Previous files had 'Lattitude' and 'Longitude' headers swapped.
    T9 onward has them correct. We detect by value range:
      latitude  for Tulsa is ~36.1  (positive, 30-40)
      longitude for Tulsa is ~-95.9 (negative, -100 to -90)
    """
    col_a_val = row.get("Lattitude")
    col_b_val = row.get("Longitude")

    try:
        a = float(col_a_val)
        b = float(col_b_val)
    except (TypeError, ValueError):
        return None, None

    # Identify which value is latitude (~36) and which is longitude (~-96)
    if 30 <= a <= 40 and -105 <= b <= -85:
        return a, b   # correctly labeled
    elif 30 <= b <= 40 and -105 <= a <= -85:
        return b, a   # swapped labels — correct silently
    else:
        return None, None  # unrecognizable — flag as missing

def process_poles(wb):
    """Convert a data sheet into structured item list.
    Detects the data sheet name automatically — any sheet not named '*Definition'."""
    data_sheet = next(
        (s for s in wb.sheetnames if not s.endswith('Definition')), None
    )
    if not data_sheet:
        print(f"  WARNING: No data sheet found. Sheets: {wb.sheetnames}")
        return []

    rows = parse_sheet(wb, data_sheet)
    poles = []
    for r in rows:
        lat, lng = extract_coords(r)

        pole = {
            "tagId":          str_or_null(r.get("Tag #")),
            "deviceNum":      str_or_null(r.get("Device #")),
            "type":           str_or_null(r.get("Type")),
            "description":    str_or_null(r.get("Description")),
            "locationId":     str_or_null(r.get("Location ID")),
            "subType":        str_or_null(r.get("Sub Type Name")),
            "lat":            lat,
            "lng":            lng,
            "electrical": {
                "panel":          str_or_null(r.get("Panel")),
                "circuitBreaker": str_or_null(r.get("Circuit Breaker")),
                "dimmerModule":   int_or_null(r.get("Dimmer Module")),
                "dimmerChannel":  str_or_null(r.get("Dimmer Channel")),
                "controlLocation":str_or_null(r.get("Control Location")),
                "controlLabel":   str_or_null(r.get("Control Label"))
            }
        }
        poles.append(pole)

    poles.sort(key=lambda p: p["deviceNum"] or "")
    return poles

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    poles = []
    missing = []

    for ptype, fname in sorted(EXCEL_SOURCES.items()):
        if not os.path.exists(fname):
            print(f"WARNING: {fname} not found — skipping {ptype}")
            missing.append(fname)
            continue
        print(f"Reading {fname}...")
        wb = load_workbook(fname, read_only=True)
        batch = process_poles(wb)
        print(f"  Processed {len(batch)} {ptype} items")
        poles.extend(batch)

    if missing:
        print(f"\nNote: {len(missing)} source file(s) were not found and were skipped.")

    # Sort all poles together by type then device number
    poles.sort(key=lambda p: (p["type"], p["deviceNum"] or ""))

    output = {
        "meta": {
            "version": "1.0",
            "generatedFrom": list(EXCEL_SOURCES.values()),
            "description": "GGP Lighting Assessment Tool — Pole Configuration"
        },
        "mapConfig":        MAP_CONFIG,
        "conditionScale":   CONDITION_SCALE,
        "poleTypes":        POLE_TYPE_DEFINITIONS,
        "poles":            poles
    }

    os.makedirs("data", exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2)

    print(f"  Written to {OUTPUT_FILE}")
    print("Done.")

if __name__ == "__main__":
    main()
